from PIL import Image
import webcolors as wb
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

xls = Workbook()
sheet = xls.active

image = Image.open("./images/Beelzebub.600.418332.jpg").convert("RGB")
image_size = image.size
pixel = image.load()
for i in range(image_size[0]):
    sheet.column_dimensions[f"{get_column_letter(i+1)}"].width = 2.5
    for j in range(image_size[1]):
        hex_val = wb.rgb_to_hex(pixel[i,j])
        sheet[f"{get_column_letter(i+1)}{j+1}"].fill = PatternFill("solid", fgColor=f"{hex_val[1:]}")

xls.save("image.xlsx")